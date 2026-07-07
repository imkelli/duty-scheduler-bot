"""
Генератор демо-артефактов на полностью вымышленных данных:

  1. docs/demo_schedule.png   — скриншот PNG-графика для README
     (использует боевой рендерер image_render.render_schedule_png)
  2. examples/schedule_demo.xlsx — демо-Excel для локального запуска бота
     (лист-год с расписанием + лист Phones с контактами)

Запуск из корня репозитория:  python docs/generate_demo.py
Реальные данные не используются и не затрагиваются.
"""
import os
import sys
from datetime import date, timedelta

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.services import image_render

# ── Вымышленные инженеры и подразделения ────────────────────────────────────
DEMO_ENGINEERS = [
    # (ФИО, телефон, тег, email)
    ("Иванов Иван",     "+7 900 111-22-33", "@ivanov_demo",   "ivanov@example.com"),
    ("Петров Пётр",     "+7 900 222-33-44", "@petrov_demo",   "petrov@example.com"),
    ("Сидоров Сидор",   "+7 900 333-44-55", "@sidorov_demo",  ""),
    ("Кузнецов Кузьма", "+7 900 444-55-66", "@kuznetsov_demo", "kuznetsov@example.com"),
]

DEMO_PROJECTS = {
    "Иванов Иван":     ["Отдел А · Портал заявок", "Отдел А · CRM"],
    "Петров Пётр":     ["Отдел А · Мониторинг", "Отдел Б · Биллинг", "Отдел Б · Склад"],
    "Сидоров Сидор":   ["Отдел Б · Отчётность"],
    "Кузнецов Кузьма": ["Отдел А · Шина данных", "Отдел Б · Личный кабинет"],
}

# Проект, оставшийся без дежурного, — показывает красную строку в графике
DEMO_ORPHAN = "Отдел Б · Резервное копирование"


def _demo_period() -> str:
    """Ближайшая неделя пн–вс в формате 'DD.MM - DD.MM'."""
    today = date.today()
    monday = today - timedelta(days=today.weekday())
    sunday = monday + timedelta(days=6)
    return f"{monday:%d.%m} - {sunday:%d.%m}"


def make_png(period: str, out_path: str):
    rows = []
    for name, phone, tag, email in DEMO_ENGINEERS:
        rows.append({
            "full_name": name,
            "phone": phone,
            "telegram_tag": tag,
            "email": email,
            "projects": DEMO_PROJECTS[name],
            "no_duty": False,
        })
    rows.append({
        "full_name": "БЕЗ ДЕЖУРНОГО",
        "phone": "",
        "telegram_tag": "",
        "email": "",
        "projects": [DEMO_ORPHAN],
        "no_duty": True,
    })
    image_render.render_schedule_png(period, rows, out_path)
    print(f"PNG:  {out_path}")


def make_xlsx(period: str, out_path: str):
    from openpyxl import Workbook
    wb = Workbook()

    # Лист расписания: имя листа = текущий год, периоды в строке 1 (колонки B+),
    # проекты в колонке A (строки 3–4 у бота служебные — оставляем пустыми).
    ws = wb.active
    ws.title = str(date.today().year)
    ws["B1"] = period
    all_projects, owners = [], []
    for name, projects in DEMO_PROJECTS.items():
        for p in projects:
            all_projects.append(p)
            owners.append(name)
    all_projects.append(DEMO_ORPHAN)
    owners.append("")  # без дежурного
    row = 5  # начинаем с 5-й: строки 3–4 бот пропускает как служебные
    for proj, owner in zip(all_projects, owners):
        ws.cell(row=row, column=1, value=proj)
        if owner:
            ws.cell(row=row, column=2, value=owner)
        row += 1

    # Лист Phones: A=ФИО, B=телефон (+email в той же ячейке), C=тег, D=роль
    ph = wb.create_sheet("Phones")
    ph.append(["ФИО", "Телефон", "Telegram", "Роль"])
    for name, phone, tag, email in DEMO_ENGINEERS:
        cell_b = f"{phone} {email}".strip()
        ph.append([name, cell_b, tag, "Инженер"])

    wb.save(out_path)
    print(f"XLSX: {out_path}")


if __name__ == "__main__":
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    period = _demo_period()
    make_png(period, os.path.join(base, "docs", "demo_schedule.png"))
    make_xlsx(period, os.path.join(base, "examples", "schedule_demo.xlsx"))
