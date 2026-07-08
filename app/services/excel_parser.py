import os
import re
from datetime import date, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.exceptions import InvalidFileException
from typing import Optional
from app.db import database


class ExcelError(Exception):
    """Excel-related error with a user-facing message in Russian."""


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


# ─── Workbook cache (invalidated by mtime) ───────────────────────────────────
_workbook_cache: dict[str, tuple[float, "object"]] = {}


def _load_cached(path: str):
    """Return the parsed workbook, reusing the cached copy if mtime hasn't changed."""
    mtime = os.path.getmtime(path)
    cached = _workbook_cache.get(path)
    if cached and cached[0] == mtime:
        return cached[1]
    wb = load_workbook(path, data_only=True)
    _workbook_cache[path] = (mtime, wb)
    return wb


def _open_excel(path: str):
    """
    Open the Excel file and return the workbook, translating low-level errors
    into ExcelError with user-facing Russian messages.
    """
    filename = os.path.basename(path)
    if not os.path.exists(path):
        raise ExcelError(f"Файл <code>{filename}</code> не найден")
    try:
        return _load_cached(path)
    except PermissionError:
        # File is locked — typically opened in Excel
        # Invalidate cache so next attempt re-reads after user closes Excel
        invalidate_cache(path)
        raise ExcelError(
            f"Файл <code>{filename}</code> открыт в Excel — "
            f"закройте его и попробуйте снова"
        )
    except InvalidFileException:
        invalidate_cache(path)
        raise ExcelError(f"Структура файла <code>{filename}</code> повреждена")
    except FileNotFoundError:
        raise ExcelError(f"Файл <code>{filename}</code> не найден")
    except OSError as e:
        invalidate_cache(path)
        # Windows sharing violation when file is open in Excel: winerror 32, 33
        msg = str(e).lower()
        winerr = getattr(e, "winerror", None)
        is_lock = (
            winerr in (32, 33)
            or "being used" in msg
            or "sharing" in msg
            or "process cannot" in msg
        )
        if is_lock:
            raise ExcelError(
                f"Файл <code>{filename}</code> занят другим процессом — "
                f"закройте его и попробуйте снова"
            )
        raise ExcelError(f"Не удалось открыть <code>{filename}</code>: {type(e).__name__}")
    except Exception as e:
        invalidate_cache(path)
        raise ExcelError(f"Структура файла <code>{filename}</code> повреждена ({type(e).__name__})")


def _require_sheet(wb, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        raise ExcelError(f"В файле отсутствует лист <code>{sheet_name}</code>")
    return wb[sheet_name]


def invalidate_cache(path: Optional[str] = None):
    if path is None:
        _workbook_cache.clear()
    else:
        _workbook_cache.pop(path, None)


_EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+\.[\w.-]+")


def _split_phone_email(raw: str) -> tuple[str, str]:
    """
    Column B may contain phone and email together (separated by space, line break,
    or Excel's _x000D_ artifact). Extract email if present; phone is what's left.
    """
    if not raw:
        return "", ""
    text = raw.replace("_x000D_", " ").replace("\r", " ").replace("\n", " ")
    m = _EMAIL_RE.search(text)
    if m:
        email = m.group(0)
        phone = (text[:m.start()] + text[m.end():]).strip()
    else:
        email = ""
        phone = text.strip()
    phone = re.sub(r"\s+", " ", phone).strip()
    return phone, email


async def import_phones(excel_path: str) -> tuple[int, list[str]]:
    """
    Import engineers from the Phones sheet into the database atomically.
    Validates the file first, parses everything in memory, then commits in a
    single transaction. On any error, the DB is left untouched.

    Returns (upserted_count, ambiguous_names): имена, по которым в базе
    несколько записей (тёзки), пропущены — их показывает администратору
    вызывающий код (_do_import).
    """
    from app.middlewares import security
    security.validate_excel(excel_path, required_sheet="Phones")

    # Force fresh read on import (file might have been replaced)
    invalidate_cache(excel_path)
    wb = _open_excel(excel_path)
    ws = _require_sheet(wb, "Phones")
    pending: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        full_name = _clean(row[0] if len(row) > 0 else "")
        raw_phone = _clean(row[1] if len(row) > 1 else "")
        tag = _clean(row[2] if len(row) > 2 else "")
        if not full_name or full_name == "-":
            continue
        if tag == "-":
            tag = ""
        phone, email = _split_phone_email(raw_phone)
        # Cap field lengths defensively (DB has no length constraints)
        pending.append({
            "full_name": full_name[:200],
            "phone":     phone[:200],
            "telegram_tag": tag[:64],
            "email":     email[:200],
        })
    wb.close()

    if not pending:
        raise ValueError("В листе 'Phones' не найдено ни одной валидной строки.")

    return await database.bulk_upsert_engineers(pending)


def _parse_period_start(label: str) -> Optional[date]:
    """
    Parse the start date from a period label like '06.04 - 12.04' or '07.04 - 13. 04'.
    Year is always the current calendar year.
    """
    m = re.match(r"(\d{1,2})\s*\.\s*(\d{1,2})", label.strip())
    if not m:
        return None
    day, month = int(m.group(1)), int(m.group(2))
    try:
        return date(date.today().year, month, day)
    except ValueError:
        return None


def get_periods(excel_path: str, filter_weeks: Optional[int] = None) -> list[tuple[int, str]]:
    """
    Return periods from row 1 of the schedule sheet, sorted by start date ascending.
    If filter_weeks is set, only return periods whose start date falls within
    [Monday of current week, Monday of current week + filter_weeks weeks].
    Duplicate start dates (same week in different sheet sections) are deduplicated,
    keeping the first occurrence by column order.
    """
    wb = _open_excel(excel_path)
    sheet_name = _find_schedule_sheet(wb)
    if sheet_name is None or sheet_name not in wb.sheetnames:
        raise ExcelError("В файле отсутствует лист с расписанием (год, например <code>2025</code>)")
    ws = wb[sheet_name]

    today = date.today()
    week_start = today - timedelta(days=today.weekday())  # Monday of current week
    week_end = week_start + timedelta(weeks=filter_weeks) if filter_weeks is not None else None

    # iso_week key -> (start_date, col, label): keep earliest start per week
    best: dict[tuple[int, int], tuple[date, int, str]] = {}

    for col in ws.iter_cols(min_col=2, min_row=1, max_row=1):
        cell = col[0]
        label = _clean(cell.value)
        if not label:
            continue
        start = _parse_period_start(label)
        if start is None:
            continue
        if week_end is not None and not (week_start <= start <= week_end):
            continue
        iso_week = (start.isocalendar()[0], start.isocalendar()[1])  # (year, week_num)
        if iso_week not in best or start < best[iso_week][0]:
            best[iso_week] = (start, cell.column, label)

    periods = sorted(best.values(), key=lambda x: x[0])
    return [(col, label) for _, col, label in periods]


def get_user_duties(excel_path: str, full_name: str, filter_weeks: int = 3) -> list[tuple[str, list[str]]]:
    """
    For a given engineer, return [(period_label, [projects])] for the next N weeks
    (inclusive of the current week). Empty list of projects means no duty that week.
    """
    periods = get_periods(excel_path, filter_weeks=filter_weeks)
    result: list[tuple[str, list[str]]] = []
    for col, label in periods:
        duty_map = get_duty_map(excel_path, col)
        result.append((label, duty_map.get(full_name, [])))
    return result


def get_duty_map(excel_path: str, col_index: int) -> dict[str, list[str]]:
    """
    Read the given column and return {full_name: [project1, project2, ...]}
    Rows 3 and 4 (index 3,4) are skipped (service rows).
    Names in cells can be slash-separated.
    """
    wb = _open_excel(excel_path)
    sheet_name = _find_schedule_sheet(wb)
    if sheet_name is None or sheet_name not in wb.sheetnames:
        raise ExcelError("В файле отсутствует лист с расписанием (год, например <code>2025</code>)")
    ws = wb[sheet_name]
    duty_map: dict[str, list[str]] = {}

    for row in ws.iter_rows(min_row=2, values_only=False):
        row_num = row[0].row
        if row_num in (3, 4):
            continue
        project_cell = row[0]
        project = _clean(project_cell.value)
        if not project:
            continue
        duty_cell = ws.cell(row=row_num, column=col_index)
        duty_val = _clean(duty_cell.value)
        if not duty_val or duty_val == "-":
            continue
        names = [n.strip() for n in duty_val.split("/") if n.strip()]
        for name in names:
            duty_map.setdefault(name, []).append(project)

    return duty_map


def _find_schedule_sheet(wb) -> str:
    # Prefer sheet with the highest numeric year (e.g. '2025' over '2024')
    year_sheets = [(int(name.strip()), name) for name in wb.sheetnames if name.strip().isdigit()]
    if year_sheets:
        return max(year_sheets)[1]
    # fallback: first sheet that is not Phones
    for name in wb.sheetnames:
        if name.lower() != "phones":
            return name
    return wb.sheetnames[0]


def generate_schedule_xlsx(
    period: str,
    assignments: list[dict],
    output_path: Optional[str] = None,
) -> str:
    """
    Generate a schedule .xlsx file.
    Each assignment dict must have keys:
      full_name, phone, telegram_tag, projects (list of str)
    Returns the path to the generated file.
    """
    if output_path is None:
        safe_period = period.replace(".", "-").replace(" ", "_").replace("/", "-")
        output_path = f"schedule_{safe_period}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "График дежурств"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    has_status = bool(assignments) and "status" in assignments[0]
    headers = ["Имя и фамилия", "Телефон", "Telegram", "Проекты"]
    col_widths = [30, 30, 22, 50]
    if has_status:
        headers.append("Статус")
        col_widths.append(28)

    for col_num, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 20

    def _meaningful(value) -> str:
        """Return the trimmed value, or empty string if it represents 'no value'."""
        if value is None:
            return ""
        s = str(value).strip()
        if s in ("", "-", "—", "–"):
            return ""
        return s

    red_fill = PatternFill("solid", fgColor="FFC7CE")          # light red
    red_font = Font(bold=True, color="9C0006")                  # dark red text

    for row_num, a in enumerate(assignments, start=2):
        projects_str = " │ ".join(a["projects"])
        phone = _meaningful(a.get("phone"))
        email = _meaningful(a.get("email"))
        if phone and email:
            phone_email = f"{phone}\n{email}"
            phone_lines = 2
        else:
            phone_email = phone or email
            phone_lines = 1
        no_duty = bool(a.get("no_duty"))
        name_value = a["full_name"]
        values = [name_value, phone_email, a["telegram_tag"], projects_str]
        if has_status:
            values.append(a.get("status", ""))
        for col_num, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=val)
            cell.border = border
            cell.alignment = wrap
            if no_duty:
                cell.fill = red_fill
                if col_num == 1:
                    cell.font = red_font
        lines_in_row = max(len(a["projects"]), phone_lines)
        ws.row_dimensions[row_num].height = max(20, 16 * lines_in_row)

    ws.freeze_panes = "A2"
    wb.save(output_path)
    return output_path
