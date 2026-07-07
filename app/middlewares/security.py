"""
Security utilities for the duty bot:
  - Whitelist + admin authorization middleware
  - Per-user rate limiting middleware with temporary bans
  - HTML/text input sanitization helpers
  - User-id masking for logs
  - DB backup rotation
  - Excel file validation
All security events are written to security.log; tokens, phones, emails and
full user_ids are never logged in plaintext.
"""
from __future__ import annotations

import logging
import os
import shutil
import time
from collections import defaultdict, deque
from datetime import datetime
from html import escape
from pathlib import Path
from typing import Any, Awaitable, Callable, Optional

from aiogram import BaseMiddleware
from aiogram.types import CallbackQuery, Message, TelegramObject, User

from app.db import database

# ─── Limits ──────────────────────────────────────────────────────────────────
MAX_INPUT_LENGTH       = 100
MAX_MSG_PER_MIN        = 20
MAX_CMD_PER_SEC        = 5
BAN_DURATION_SEC       = 5 * 60
MAX_EXCEL_SIZE_BYTES   = 10 * 1024 * 1024
MAX_BACKUPS            = 10
BACKUP_DIR             = Path("backups")

# ─── Logger ──────────────────────────────────────────────────────────────────
_security_logger: Optional[logging.Logger] = None


def get_logger() -> logging.Logger:
    global _security_logger
    if _security_logger is not None:
        return _security_logger
    lg = logging.getLogger("duty_bot.security")
    lg.setLevel(logging.INFO)
    lg.propagate = False
    fh = logging.FileHandler("security.log", encoding="utf-8")
    fh.setFormatter(logging.Formatter(
        "%(asctime)s | %(levelname)-7s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    ))
    lg.addHandler(fh)
    _security_logger = lg
    return lg


# ─── Helpers ─────────────────────────────────────────────────────────────────
def mask_user_id(user_id: Optional[int]) -> str:
    if user_id is None:
        return "?"
    s = str(user_id)
    return "***" + s[-4:] if len(s) >= 4 else "***"


def html_safe(value: Any) -> str:
    """Escape HTML special chars; safe for use inside HTML messages."""
    return escape("" if value is None else str(value), quote=False)


def sanitize_text(text: Optional[str], max_length: int = MAX_INPUT_LENGTH) -> str:
    """Strip control characters, collapse whitespace, enforce length cap."""
    if not text:
        return ""
    cleaned_chars = []
    for ch in text:
        if ch in ("\n", "\t", " "):
            cleaned_chars.append(ch)
        elif ch.isprintable():
            cleaned_chars.append(ch)
    cleaned = "".join(cleaned_chars).strip()
    return cleaned[:max_length]


# ─── Rate limiting state (in-memory) ─────────────────────────────────────────
_msg_window: dict[int, deque[float]] = defaultdict(lambda: deque(maxlen=64))
_cmd_window: dict[int, deque[float]] = defaultdict(lambda: deque(maxlen=16))
_bans: dict[int, float] = {}


def _is_banned(user_id: int) -> tuple[bool, int]:
    until = _bans.get(user_id, 0.0)
    remaining = int(until - time.time())
    return (remaining > 0, max(remaining, 0))


def _record_and_check(user_id: int, is_command: bool) -> Optional[str]:
    """Record event and return reason string if rate limit exceeded, else None."""
    now = time.time()

    msg_q = _msg_window[user_id]
    msg_q.append(now)
    while msg_q and msg_q[0] < now - 60:
        msg_q.popleft()
    if len(msg_q) > MAX_MSG_PER_MIN:
        _bans[user_id] = now + BAN_DURATION_SEC
        return "msg_per_min"

    if is_command:
        cmd_q = _cmd_window[user_id]
        cmd_q.append(now)
        while cmd_q and cmd_q[0] < now - 1:
            cmd_q.popleft()
        if len(cmd_q) > MAX_CMD_PER_SEC:
            _bans[user_id] = now + BAN_DURATION_SEC
            return "cmd_per_sec"
    return None


def _extract_user(event: TelegramObject) -> Optional[User]:
    if isinstance(event, (Message, CallbackQuery)):
        return event.from_user
    return None


async def _reply_blocked(event: TelegramObject, html: str, plain: str):
    try:
        if isinstance(event, CallbackQuery):
            await event.answer(plain, show_alert=True)
        elif isinstance(event, Message):
            await event.answer(html)
    except Exception:
        pass


# ─── Rate limit middleware ───────────────────────────────────────────────────
class RateLimitMiddleware(BaseMiddleware):
    async def __call__(
        self,
        handler: Callable[[TelegramObject, dict[str, Any]], Awaitable[Any]],
        event: TelegramObject,
        data: dict[str, Any],
    ) -> Any:
        user = _extract_user(event)
        if user is None:
            return await handler(event, data)
        uid = user.id

        banned, remaining = _is_banned(uid)
        if banned:
            get_logger().warning(
                f"BAN_ACTIVE user={mask_user_id(uid)} remaining={remaining}s"
            )
            await _reply_blocked(
                event,
                f"<b>Доступ временно заблокирован</b>\n<i>Осталось: {remaining} сек.</i>",
                f"Доступ заблокирован. Осталось: {remaining} сек.",
            )
            return

        is_cmd = isinstance(event, Message) and (event.text or "").startswith("/")
        reason = _record_and_check(uid, is_cmd)
        if reason:
            get_logger().warning(
                f"RATE_LIMIT user={mask_user_id(uid)} reason={reason} -> ban {BAN_DURATION_SEC}s"
            )
            await _reply_blocked(
                event,
                "<b>Слишком много запросов</b>\n<i>Доступ заблокирован на 5 минут.</i>",
                "Слишком много запросов. Бан на 5 минут.",
            )
            return

        return await handler(event, data)


# ─── Cached admin Telegram tag ───────────────────────────────────────────────
_admin_tag: str = ""  # "@tag" or "" when unknown


def set_admin_tag(tag: Optional[str]):
    global _admin_tag
    t = (tag or "").strip()
    if t and not t.startswith("@"):
        t = "@" + t
    _admin_tag = t


def admin_tag() -> str:
    """Raw '@tag' or '' when unknown."""
    return _admin_tag


def admin_mention() -> str:
    """Standalone reference: '@tag' if known, else 'администратору'."""
    return _admin_tag or "администратору"


def admin_tag_suffix() -> str:
    """' @tag' to append after the word 'администратор(у)', or '' when unknown."""
    return f" {_admin_tag}" if _admin_tag else ""


# ─── Link-flow allowlist ─────────────────────────────────────────────────────
# Users currently going through the /start account-linking search are allowed
# to send messages / press linkreq buttons even though they are not yet linked.
_link_flow_users: set[int] = set()


def add_link_flow_user(user_id: int):
    _link_flow_users.add(user_id)


def remove_link_flow_user(user_id: int):
    _link_flow_users.discard(user_id)


# Callback prefixes always allowed for unlinked users (the linking flow itself)
_LINK_FLOW_CALLBACKS = ("linkreq_pick:", "linkreq_cancel")


# ─── Authorization middleware (binding-based access control) ──────────────────
class AuthMiddleware(BaseMiddleware):
    def __init__(self, admin_id: int):
        self.admin_id = admin_id

    async def __call__(
        self,
        handler: Callable[[TelegramObject, dict[str, Any]], Awaitable[Any]],
        event: TelegramObject,
        data: dict[str, Any],
    ) -> Any:
        user = _extract_user(event)
        if user is None:
            return await handler(event, data)
        uid = user.id

        # Admin always has full access (identified by ADMIN_ID, not by a DB record)
        if uid == self.admin_id:
            return await handler(event, data)

        # /start — entry point, always allowed
        text = (event.text or "") if isinstance(event, Message) else ""
        if text.startswith("/start"):
            return await handler(event, data)

        # Linking-flow callbacks — allowed even for unlinked users
        cb_data = event.data if isinstance(event, CallbackQuery) else ""
        if any(cb_data.startswith(p) for p in _LINK_FLOW_CALLBACKS):
            return await handler(event, data)

        # Users mid-way through the link search (typing their tag) — allowed
        if uid in _link_flow_users:
            return await handler(event, data)

        # Binding-based gate
        eng = await database.get_engineer_by_user_id(uid)
        pending = await database.get_pending_request_by_user(uid)

        if eng and not pending:
            return await handler(event, data)  # fully linked, no open request

        # Blocked — choose the right message
        access_line = f"\n\n<i>По вопросам доступа: {admin_mention()}</i>"
        if pending:
            html = (
                "<b>Ожидание подтверждения</b>\n"
                "\n"
                "<i>Ваш запрос рассматривается администратором.</i>"
            )
            plain = "Ваш запрос рассматривается администратором."
        else:
            last = await database.get_last_resolved_request(uid)
            was_unlinked = bool(
                last and last["request_type"] == "unlink" and last["status"] == "approved"
            )
            if was_unlinked:
                html = (
                    "<b>Доступ ограничен</b>\n"
                    "\n"
                    "<i>Ваш аккаунт отвязан. Для возобновления работы введите /start.</i>"
                    + access_line
                )
            else:
                html = (
                    "<b>Доступ ограничен</b>\n"
                    "\n"
                    "<i>Для использования бота необходимо привязать аккаунт. "
                    "Введите /start для начала.</i>"
                    + access_line
                )
            plain = "Доступ ограничен. Введите /start."

        get_logger().info(
            f"ACCESS_BLOCKED user={mask_user_id(uid)} "
            f"reason={'pending' if pending else 'no_binding'}"
        )
        await _reply_blocked(event, html, plain)
        return


# ─── Admin enforcement (defence in depth, even if button hidden) ─────────────
def admin_only(admin_id: int, user_id: int, action: str) -> bool:
    """Returns True if allowed. Logs every attempt."""
    allowed = (user_id == admin_id)
    get_logger().info(
        f"ADMIN_CHECK action={action} user={mask_user_id(user_id)} "
        f"result={'ALLOW' if allowed else 'DENY'}"
    )
    return allowed


# ─── Replacement chain protection ────────────────────────────────────────────
def validate_replacement(
    chain: list[dict],
    proposer_engineer_id: Optional[int],
    original_engineer_id: int,
    candidate_engineer_id: int,
) -> Optional[str]:
    """
    Returns an error string explaining why the replacement is rejected,
    or None if the replacement is acceptable.
    """
    if candidate_engineer_id == proposer_engineer_id:
        return "Нельзя предложить заменой самого себя."
    if candidate_engineer_id == original_engineer_id:
        return "Нельзя предложить заменой исходного дежурного."
    seen_ids = {step.get("engineer_id") for step in chain}
    if candidate_engineer_id in seen_ids:
        return "Этот человек уже участвовал в цепочке замен."
    return None


def log_replacement(session_id: int, assignment_id: int, chain: list[dict]):
    """Append a snapshot of the replacement chain with current timestamp."""
    chain_str = " → ".join(str(step.get("engineer_id")) for step in chain)
    get_logger().info(
        f"REPLACEMENT session={session_id} assignment={assignment_id} chain={chain_str}"
    )


# ─── DB backup with rotation ─────────────────────────────────────────────────
def backup_database(db_path: str = "duty_bot.db") -> Optional[Path]:
    """Сделать копию БД в backups/ с таймстампом; ротация старых копий."""
    if not os.path.exists(db_path):
        return None
    BACKUP_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = BACKUP_DIR / f"{Path(db_path).stem}_{ts}.db"
    shutil.copy2(db_path, dst)
    get_logger().info(f"BACKUP created file={dst.name}")
    _rotate_backups(Path(db_path).stem)
    return dst


def _rotate_backups(stem: str):
    backups = sorted(
        BACKUP_DIR.glob(f"{stem}_*.db"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    for old in backups[MAX_BACKUPS:]:
        try:
            old.unlink()
            get_logger().info(f"BACKUP rotated_out file={old.name}")
        except OSError:
            pass


# ─── Excel validation ────────────────────────────────────────────────────────
def validate_excel(path: str, required_sheet: str = "Phones"):
    """Raise excel_parser.ExcelError on any validation failure."""
    from app.services import excel_parser  # local import to avoid circular deps
    filename = os.path.basename(path)
    if not os.path.exists(path):
        raise excel_parser.ExcelError(f"Файл <code>{filename}</code> не найден")
    size = os.path.getsize(path)
    if size > MAX_EXCEL_SIZE_BYTES:
        raise excel_parser.ExcelError(
            f"Файл слишком большой ({size // (1024 * 1024)} МБ, лимит "
            f"{MAX_EXCEL_SIZE_BYTES // (1024 * 1024)} МБ)"
        )
    if size == 0:
        raise excel_parser.ExcelError(f"Файл <code>{filename}</code> пуст")

    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except PermissionError:
        raise excel_parser.ExcelError(
            f"Файл <code>{filename}</code> открыт в Excel — закройте его и попробуйте снова"
        )
    except InvalidFileException:
        raise excel_parser.ExcelError(f"Структура файла <code>{filename}</code> повреждена")
    except OSError as e:
        winerr = getattr(e, "winerror", None)
        if winerr in (32, 33):
            raise excel_parser.ExcelError(
                f"Файл <code>{filename}</code> занят другим процессом — закройте его и попробуйте снова"
            )
        raise excel_parser.ExcelError(f"Не удалось открыть <code>{filename}</code>: {type(e).__name__}")
    except Exception as e:
        raise excel_parser.ExcelError(
            f"Структура файла <code>{filename}</code> повреждена ({type(e).__name__})"
        )

    try:
        if required_sheet not in wb.sheetnames:
            raise excel_parser.ExcelError(f"В файле отсутствует лист <code>{required_sheet}</code>")
        ws = wb[required_sheet]
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not first_row or sum(1 for v in first_row if v) < 3:
            raise excel_parser.ExcelError(
                f"Лист <code>{required_sheet}</code> должен содержать минимум 3 заполненные колонки в шапке"
            )
    finally:
        wb.close()
